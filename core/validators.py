"""
Input validation functions for the Nessus Report Generator.

Handles validation of CSV files, Excel files, filenames, and directories.
All functions are pure (no side effects) and return (is_valid, error_message).
"""

import logging
import os
from pathlib import Path
from typing import Optional

from config import INVALID_FILENAME_CHARS, MAX_FILENAME_LENGTH

logger = logging.getLogger(__name__)


# =============================================================================
# Exceptions
# =============================================================================

class ValidationError(Exception):
    """Custom exception for validation errors with user-friendly messages."""

    def __init__(self, message: str, details: str = "") -> None:
        self.message = message
        self.details = details
        super().__init__(message)


# =============================================================================
# CSV validation
# =============================================================================

def validate_csv_file(filepath: str) -> tuple[bool, str]:
    """
    Validate that *filepath* points to a readable, non-empty .csv file.

    Returns:
        (is_valid, error_message)
    """
    try:
        path = Path(filepath)

        if not path.exists():
            return False, f"File not found: {path.name}"

        if path.suffix.lower() != ".csv":
            return False, f"Invalid file type: {path.suffix}. Expected .csv"

        if not path.is_file():
            return False, f"Path is not a file: {path.name}"

        return _check_file_readable(path)

    except Exception as exc:
        return False, f"Validation error: {exc}"


def _check_file_readable(path: Path) -> tuple[bool, str]:
    """Try to read the first line with utf-8, then cp1252."""
    for enc in ("utf-8", "cp1252"):
        try:
            with open(path, "r", encoding=enc) as fh:
                first_line = fh.readline()
                if not first_line.strip():
                    return False, f"File appears to be empty: {path.name}"
                return True, ""
        except UnicodeDecodeError:
            continue
        except PermissionError:
            return False, f"Permission denied: {path.name}"
        except Exception:
            return False, f"Cannot read file: {path.name}"

    return False, f"Cannot read file (encoding issue): {path.name}"


def validate_csv_columns(
    headers: list[str],
    filepath: str,
) -> tuple[bool, list[str]]:
    """
    Validate that a CSV header row has the expected Nessus columns.

    Args:
        headers:  List of header strings from the CSV.
        filepath: Path (for error messages only).

    Returns:
        (is_valid, list_of_warnings)
    """
    from core.csv_processor import normalize_column_name

    warnings: list[str] = []
    headers_lower = [h.lower().strip() for h in headers]
    filename = Path(filepath).name

    ip_variants = {"ip address", "host", "ip", "target", "hostname"}
    has_ip = bool(ip_variants & set(headers_lower))
    if not has_ip:
        warnings.append(f"Warning: '{filename}' may be missing IP Address column")

    name_variants = {"plugin name", "name", "issue name", "vulnerability", "title"}
    has_name = bool(name_variants & set(headers_lower))
    if not has_name:
        warnings.append(f"Warning: '{filename}' may be missing Plugin Name column")

    # Detect columns that will collapse under normalization
    norm_counts: dict[str, list[str]] = {}
    for h in headers:
        norm = normalize_column_name(h)
        norm_counts.setdefault(norm, []).append(h)

    for norm, cols in norm_counts.items():
        if len(cols) > 1:
            warnings.append(
                f"Note: Multiple columns map to '{norm}' "
                f"({', '.join(cols)}). Only one will be kept."
            )

    is_valid = has_ip or has_name
    return is_valid, warnings


# =============================================================================
# Excel validation (for rescan mode)
# =============================================================================

def validate_excel_file(filepath: str) -> tuple[bool, str]:
    """
    Validate a previous Excel report for rescan mode.

    Returns:
        (is_valid, error_message)
    """
    try:
        path = Path(filepath)

        if not path.exists():
            return False, f"Previous Excel file not found: {path.name}"

        if path.suffix.lower() != ".xlsx":
            return False, f"Invalid file type: {path.suffix}. Expected .xlsx"

        if not path.is_file():
            return False, f"Path is not a file: {path.name}"

        return _check_excel_contents(filepath)

    except Exception as exc:
        return False, f"Validation error: {exc}"


def _check_excel_contents(filepath: str) -> tuple[bool, str]:
    """Open the workbook and verify it has a valid Findings sheet."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)

        if "Findings" not in wb.sheetnames:
            wb.close()
            return False, "Excel file is missing 'Findings' sheet"

        ws = wb["Findings"]
        raw_headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()

        headers_lower = [str(h).lower() for h in raw_headers if h]
        required = ["ip address", "issue name", "port"]
        missing = [r.title() for r in required if not any(r in h for h in headers_lower)]

        if missing:
            return False, f"Findings sheet missing columns: {', '.join(missing)}"

        return True, ""

    except PermissionError:
        return False, f"Permission denied: {Path(filepath).name}"
    except Exception as exc:
        return False, f"Cannot read Excel file: {exc}"


# =============================================================================
# Output path validation
# =============================================================================

def validate_output_filename(filename: str) -> tuple[bool, str]:
    """
    Validate output filename for invalid chars, length, and reserved names.

    Returns:
        (is_valid, error_message)
    """
    if not filename or not filename.strip():
        return False, "Filename cannot be empty"

    filename = filename.strip()

    for char in INVALID_FILENAME_CHARS:
        if char in filename:
            return False, f"Filename contains invalid character: {char}"

    if len(filename) > MAX_FILENAME_LENGTH:
        return False, f"Filename too long. Maximum {MAX_FILENAME_LENGTH} characters"

    reserved = {
        "CON", "PRN", "AUX", "NUL",
        "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
        "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
    }
    base_name = filename.rsplit(".", 1)[0].upper()
    if base_name in reserved:
        return False, f"'{filename}' is a reserved Windows filename"

    return True, ""


def validate_output_directory(directory: str) -> tuple[bool, str]:
    """
    Validate that the output directory exists and is writable.

    Returns:
        (is_valid, error_message)
    """
    try:
        path = Path(directory)

        if not path.exists():
            return False, f"Directory does not exist: {directory}"

        if not path.is_dir():
            return False, f"Path is not a directory: {directory}"

        # Check write permission
        test_file = path / ".write_test_temp"
        try:
            test_file.touch()
            test_file.unlink()
        except PermissionError:
            return False, f"No write permission for directory: {directory}"
        except Exception as exc:
            return False, f"Cannot write to directory: {exc}"

        return True, ""

    except Exception as exc:
        return False, f"Validation error: {exc}"


# =============================================================================
# Filename utilities
# =============================================================================

def sanitize_filename(filename: str) -> str:
    """Remove invalid characters and truncate if needed."""
    for char in INVALID_FILENAME_CHARS:
        filename = filename.replace(char, "_")

    if len(filename) > MAX_FILENAME_LENGTH:
        if "." in filename:
            name, ext = filename.rsplit(".", 1)
            max_len = MAX_FILENAME_LENGTH - len(ext) - 1
            filename = name[:max_len] + "." + ext
        else:
            filename = filename[:MAX_FILENAME_LENGTH]

    return filename


def ensure_xlsx_extension(filename: str) -> str:
    """Ensure filename ends with .xlsx."""
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    return filename
